from bs4 import BeautifulSoup
import re  # 正则表达式，进行文字匹配
import urllib.request, urllib.error  # 指定URL，获取网页数据
import openpyxl  # 进行excel操作


def main():
    # 自定义5个网页链接（每页25条，总共5页，共125部电影）
    custom_urls = [
        "https://movie.douban.com/top250?start=0",   # 第1页
        "https://movie.douban.com/top250?start=25",  # 第2页
        "https://movie.douban.com/top250?start=50",  # 第3页
        "https://movie.douban.com/top250?start=75",  # 第4页
        "https://movie.douban.com/top250?start=100"  # 第5页
    ]

    # 获取数据
    datalist = getdata(custom_urls)

    # 保存数据到Excel文件
    savepath = ".\\豆瓣电影Top250前125的相关信息.xlsx"  # 保存为xlsx格式
    savedata(datalist, savepath)


# 正则表达式模式
findLink = re.compile(r'<a href="(.*?)">')  # 影片详情
findImgSrc = re.compile(r'<img.*src="(.*?)"', re.S)  # 图片链接
findTitle = re.compile(r'<span class="title">(.*)</span>')  # 影片名称
findRating = re.compile(r'<span class="rating_num" property="v:average">(.*)</span>')  # 评分
findJudge = re.compile(r'<span>(\d*)人评价</span>')  # 评价人数
findInq = re.compile(r'<span class="inq">(.*)</span>')  # 概况
findBd = re.compile(r'<p class="">(.*?)</p>', re.S)  # 相关内容（导演，演员等）


# 获取数据
def getdata(url_list):
    datalist = []
    for url in url_list:  # 遍历自定义的网页链接
        html = geturl(url)  # 获取页面HTML内容
        soup = BeautifulSoup(html, "html.parser")  # 解析HTML
        for item in soup.find_all("div", class_='item'):  # 找到每部电影的item
            data = []  # 保存单部电影的所有信息
            item = str(item)  # 转换为字符串便于正则匹配

            # 提取数据
            link = re.findall(findLink, item)[0]
            data.append(link)

            imgSrc = re.findall(findImgSrc, item)[0]
            data.append(imgSrc)

            titles = re.findall(findTitle, item)
            if len(titles) == 2:
                onetitle = titles[0]
                data.append(onetitle)
                twotitle = titles[1].replace("/", "")  # 去掉斜杠
                data.append(twotitle)
            else:
                data.append(titles[0])
                data.append(" ")

            rating = re.findall(findRating, item)[0]
            data.append(rating)

            judgeNum = re.findall(findJudge, item)[0]
            data.append(judgeNum)

            inq = re.findall(findInq, item)
            if len(inq) != 0:
                inq = inq[0].replace("。", "")
                data.append(inq)
            else:
                data.append(" ")

            bd = re.findall(findBd, item)[0]
            bd = re.sub('<br(\s+)?/>(\s+)?', " ", bd)
            bd = re.sub('/', " ", bd)
            data.append(bd.strip())

            datalist.append(data)
    return datalist


# 保存数据到Excel文件
def savedata(datalist, savepath):
    workbook = openpyxl.Workbook()  # 创建新的工作簿
    worksheet = workbook.active  # 获取当前活动工作表
    worksheet.title = "豆瓣电影Top250前125的相关信息"  # 设置工作表名称

    # 写入Excel标题行
    column = ("电影详情链接", "图片链接", "影片中文名", "影片外国名", "评分", "评价数", "概况", "相关信息")
    for i in range(8):
        worksheet.cell(row=1, column=i + 1, value=column[i])  # 设置表头

    # 写入电影数据
    for i in range(len(datalist)):
        data = datalist[i]
        for j in range(8):
            worksheet.cell(row=i + 2, column=j + 1, value=data[j])  # 设置每一行的数据

    workbook.save(savepath)  # 保存到指定路径


# 爬取网页
def geturl(url):
    head = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36"
    }
    req = urllib.request.Request(url, headers=head)
    try:
        response = urllib.request.urlopen(req)
        html = response.read().decode("utf-8")
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
    return html


if __name__ == '__main__':
    main()
    print("爬取成功！！！")
